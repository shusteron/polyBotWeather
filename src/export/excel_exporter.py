from __future__ import annotations

import os
from datetime import datetime, timezone, timedelta
from typing import Any, Optional

_ISRAEL = timezone(timedelta(hours=3))  # EEST (UTC+3, Israel summer)

import numpy as np
import pandas as pd
from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..models import CalibrationRecord, Trade


class ExcelExporter:
    """
    XLSX and CSV reporting for all trade data and analytics.
    """

    @staticmethod
    def _to_il(ts: str) -> str:
        """Convert a UTC timestamp string to Israel time (UTC+3), formatted for Excel."""
        if not ts:
            return ""
        try:
            dt = datetime.fromisoformat(ts.replace("Z", "+00:00"))
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            il = dt.astimezone(_ISRAEL)
            return il.strftime("%Y-%m-%d %H:%M")
        except Exception:
            return ts

    HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True)
    WIN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    LOSS_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    def export_full_report(
        self,
        trades: list[Trade],
        skipped_trades: list[dict],
        calibration_records: list[dict],
        output_path: str,
    ) -> str:
        """
        Create a comprehensive multi-sheet XLSX workbook.
        Returns the path to the created file.
        """
        os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)

        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        self._add_trades_sheet(wb, trades)
        self._add_skipped_trades_sheet(wb, skipped_trades)
        self._add_calibration_sheet(wb, calibration_records)
        self._add_ensemble_sheet(wb, trades)
        self._add_provider_accuracy_sheet(wb, trades)
        self._add_drawdown_sheet(wb, trades)
        self._add_performance_sheet(wb, trades, calibration_records)
        self._add_stability_sheet(wb, trades)

        wb.save(output_path)
        logger.info(f"Excel report saved to {output_path}")
        return output_path

    def export_csv(self, data: pd.DataFrame, path: str) -> str:
        """Export a DataFrame to CSV."""
        os.makedirs(os.path.dirname(path) if os.path.dirname(path) else ".", exist_ok=True)
        data.to_csv(path, index=False)
        logger.info(f"CSV exported to {path}")
        return path

    # ------------------------------------------------------------------ #
    #  Sheet builders                                                      #
    # ------------------------------------------------------------------ #

    def _add_trades_sheet(self, wb: Workbook, trades: list[Trade]) -> None:
        ws = wb.create_sheet("Trades")
        headers = [
            "Market ID", "Title", "Location", "Action", "Model Prob",
            "Market Prob", "Edge", "Confidence", "Fill Price", "Size ($)",
            "Status", "Outcome", "PnL ($)", "Entry Time (IL)", "Resolution Date (IL)",
        ]
        self._write_headers(ws, headers)

        for row_idx, trade in enumerate(trades, start=2):
            ws.cell(row_idx, 1, trade.market.id)
            ws.cell(row_idx, 2, trade.market.title[:80])
            ws.cell(row_idx, 3, trade.market.location or "")
            ws.cell(row_idx, 4, trade.signal.action)
            ws.cell(row_idx, 5, round(trade.signal.model_probability, 4))
            ws.cell(row_idx, 6, round(trade.signal.market_probability, 4))
            ws.cell(row_idx, 7, round(trade.signal.edge, 4))
            ws.cell(row_idx, 8, round(trade.signal.confidence_score.total, 1))
            ws.cell(row_idx, 9, round(trade.paper_fill_price, 4))
            ws.cell(row_idx, 10, round(trade.paper_size, 2))
            ws.cell(row_idx, 11, trade.status)
            ws.cell(row_idx, 12, str(trade.outcome) if trade.outcome is not None else "")
            ws.cell(row_idx, 13, round(trade.pnl, 2) if trade.pnl is not None else "")
            ws.cell(row_idx, 14, self._to_il(trade.timestamp or ""))
            ws.cell(row_idx, 15, self._to_il(trade.market.resolution_date or ""))

            # Colour PnL row
            if trade.pnl is not None:
                fill = self.WIN_FILL if trade.pnl >= 0 else self.LOSS_FILL
                for col in range(1, 16):
                    ws.cell(row_idx, col).fill = fill

        self._auto_width(ws)

    def _add_skipped_trades_sheet(self, wb: Workbook, skipped: list[dict]) -> None:
        ws = wb.create_sheet("Skipped Trades")
        headers = [
            "Market ID", "Title", "Location", "Model Prob", "Market Prob",
            "Edge", "Confidence", "Rejection Reasons",
        ]
        self._write_headers(ws, headers)

        for row_idx, s in enumerate(skipped, start=2):
            ws.cell(row_idx, 1, s.get("market_id", ""))
            ws.cell(row_idx, 2, str(s.get("title", ""))[:80])
            ws.cell(row_idx, 3, s.get("location", ""))
            ws.cell(row_idx, 4, round(float(s.get("model_probability", 0)), 4))
            ws.cell(row_idx, 5, round(float(s.get("market_probability", 0)), 4))
            ws.cell(row_idx, 6, round(float(s.get("edge", 0)), 4))
            ws.cell(row_idx, 7, round(float(s.get("confidence", 0)), 1))
            reasons = s.get("rejection_reasons", [])
            ws.cell(row_idx, 8, " | ".join(reasons))

        self._auto_width(ws)

    def _add_calibration_sheet(self, wb: Workbook, records: list[dict]) -> None:
        ws = wb.create_sheet("Calibration Metrics")
        headers = [
            "Market ID", "Predicted Prob", "Market Prob", "Edge",
            "Confidence", "Outcome", "PnL ($)", "Resolution Date (IL)",
        ]
        self._write_headers(ws, headers)

        for row_idx, r in enumerate(records, start=2):
            ws.cell(row_idx, 1, r.get("market_id", ""))
            ws.cell(row_idx, 2, round(float(r.get("predicted_probability", 0)), 4))
            ws.cell(row_idx, 3, round(float(r.get("market_probability", 0)), 4))
            ws.cell(row_idx, 4, round(float(r.get("edge", 0)), 4))
            ws.cell(row_idx, 5, round(float(r.get("confidence", 0)), 1))
            outcome = r.get("outcome")
            ws.cell(row_idx, 6, str(outcome) if outcome is not None else "PENDING")
            pnl = r.get("pnl")
            ws.cell(row_idx, 7, round(float(pnl), 2) if pnl is not None else "")
            ws.cell(row_idx, 8, self._to_il(r.get("resolution_date", "")))

        self._auto_width(ws)

    def _add_ensemble_sheet(self, wb: Workbook, trades: list[Trade]) -> None:
        ws = wb.create_sheet("Ensemble Analysis")
        headers = [
            "Market ID", "Title", "Ensemble Mean", "Ensemble Std",
            "Confidence", "Edge", "Action",
        ]
        self._write_headers(ws, headers)

        for row_idx, trade in enumerate(trades, start=2):
            ws.cell(row_idx, 1, trade.market.id)
            ws.cell(row_idx, 2, trade.market.title[:60])
            ws.cell(row_idx, 3, "")  # ensemble mean stored in confidence breakdown
            ws.cell(row_idx, 4, "")
            ws.cell(row_idx, 5, round(trade.signal.confidence_score.total, 1))
            ws.cell(row_idx, 6, round(trade.signal.edge, 4))
            ws.cell(row_idx, 7, trade.signal.action)

        self._auto_width(ws)

    def _add_provider_accuracy_sheet(self, wb: Workbook, trades: list[Trade]) -> None:
        ws = wb.create_sheet("Provider Accuracy")
        headers = ["Metric", "Value"]
        self._write_headers(ws, headers)

        closed = [t for t in trades if t.status == "CLOSED" and t.pnl is not None]
        wins = [t for t in closed if (t.pnl or 0) > 0]

        rows = [
            ("Total Trades", len(trades)),
            ("Closed Trades", len(closed)),
            ("Wins", len(wins)),
            ("Losses", len(closed) - len(wins)),
            ("Win Rate", f"{len(wins)/len(closed):.1%}" if closed else "N/A"),
            ("Total PnL", f"${sum(t.pnl or 0 for t in closed):.2f}"),
            ("Avg Edge", f"{sum(abs(t.signal.edge) for t in trades)/len(trades):.4f}" if trades else "N/A"),
        ]
        for row_idx, (metric, value) in enumerate(rows, start=2):
            ws.cell(row_idx, 1, metric)
            ws.cell(row_idx, 2, str(value))

        self._auto_width(ws)

    def _add_drawdown_sheet(self, wb: Workbook, trades: list[Trade]) -> None:
        ws = wb.create_sheet("Drawdown Analysis")
        headers = ["Trade #", "Cumulative PnL ($)", "Running Capital ($)", "Drawdown ($)", "Drawdown (%)"]
        self._write_headers(ws, headers)

        closed = sorted(
            [t for t in trades if t.status == "CLOSED" and t.pnl is not None],
            key=lambda t: t.timestamp or "",
        )

        cumulative = 0.0
        peak = 0.0
        starting_capital = 10_000.0

        for row_idx, trade in enumerate(closed, start=2):
            pnl = trade.pnl or 0.0
            cumulative += pnl
            running = starting_capital + cumulative
            peak = max(peak, running)
            drawdown_abs = peak - running
            drawdown_pct = drawdown_abs / peak if peak > 0 else 0.0

            ws.cell(row_idx, 1, row_idx - 1)
            ws.cell(row_idx, 2, round(cumulative, 2))
            ws.cell(row_idx, 3, round(running, 2))
            ws.cell(row_idx, 4, round(drawdown_abs, 2))
            ws.cell(row_idx, 5, round(drawdown_pct, 4))

        self._auto_width(ws)

    def _add_performance_sheet(
        self, wb: Workbook, trades: list[Trade], calibration_records: list[dict]
    ) -> None:
        ws = wb.create_sheet("Performance Metrics")
        self._write_headers(ws, ["Metric", "Value"])

        closed = [t for t in trades if t.status == "CLOSED" and t.pnl is not None]
        pnls = [t.pnl for t in closed if t.pnl is not None]

        if pnls:
            mean_pnl = float(np.mean(pnls))
            std_pnl = float(np.std(pnls)) if len(pnls) > 1 else 0.0
            sharpe_like = mean_pnl / std_pnl if std_pnl > 0 else 0.0
            max_dd = 0.0
            cumulative = 0.0
            peak = 0.0
            for pnl in pnls:
                cumulative += pnl
                peak = max(peak, cumulative)
                dd = peak - cumulative
                max_dd = max(max_dd, dd)
        else:
            mean_pnl = std_pnl = sharpe_like = max_dd = 0.0

        wins = [t for t in closed if (t.pnl or 0) > 0]
        win_rate = len(wins) / len(closed) if closed else 0.0

        # Confidence buckets
        bucket_stats = self._confidence_bucket_stats(closed)

        metrics = [
            ("Total Trades Executed", len(trades)),
            ("Closed Trades", len(closed)),
            ("Win Rate", f"{win_rate:.1%}"),
            ("Total PnL", f"${sum(pnls):.2f}" if pnls else "$0.00"),
            ("Mean PnL per Trade", f"${mean_pnl:.2f}"),
            ("Std Dev PnL", f"${std_pnl:.2f}"),
            ("Sharpe-like Ratio", f"{sharpe_like:.3f}"),
            ("Max Drawdown", f"${max_dd:.2f}"),
            ("", ""),
            ("--- Confidence Buckets ---", ""),
        ]
        for bucket, stats in bucket_stats.items():
            metrics.append((f"Bucket {bucket}", stats))

        for row_idx, (metric, value) in enumerate(metrics, start=2):
            ws.cell(row_idx, 1, metric)
            ws.cell(row_idx, 2, str(value))

        self._auto_width(ws)

    def _add_stability_sheet(self, wb: Workbook, trades: list[Trade]) -> None:
        ws = wb.create_sheet("Forecast Stability")
        headers = ["Market ID", "Title", "Location", "Confidence", "Stability Score", "Action"]
        self._write_headers(ws, headers)

        for row_idx, trade in enumerate(trades, start=2):
            ws.cell(row_idx, 1, trade.market.id)
            ws.cell(row_idx, 2, trade.market.title[:60])
            ws.cell(row_idx, 3, trade.market.location or "")
            ws.cell(row_idx, 4, round(trade.signal.confidence_score.total, 1))
            stab = trade.signal.confidence_score.breakdown.get("stability", "")
            ws.cell(row_idx, 5, stab)
            ws.cell(row_idx, 6, trade.signal.action)

        self._auto_width(ws)

    # ------------------------------------------------------------------ #
    #  Helpers                                                             #
    # ------------------------------------------------------------------ #

    def _confidence_bucket_stats(self, closed_trades: list[Trade]) -> dict:
        buckets = {
            "70-80": [],
            "80-90": [],
            "90-100": [],
        }
        for trade in closed_trades:
            conf = trade.signal.confidence_score.total
            pnl = trade.pnl or 0.0
            if 70 <= conf < 80:
                buckets["70-80"].append(pnl)
            elif 80 <= conf < 90:
                buckets["80-90"].append(pnl)
            elif conf >= 90:
                buckets["90-100"].append(pnl)

        result = {}
        for bucket, pnls in buckets.items():
            if pnls:
                wins = sum(1 for p in pnls if p > 0)
                result[bucket] = f"n={len(pnls)} win_rate={wins/len(pnls):.1%} total_pnl=${sum(pnls):.2f}"
            else:
                result[bucket] = "n=0"
        return result

    def _write_headers(self, ws, headers: list[str]) -> None:
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(1, col, header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

    def _auto_width(self, ws) -> None:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)
